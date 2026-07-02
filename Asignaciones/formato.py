from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ======================================
# LIBRO
# ======================================

wb = Workbook()
ws = wb.active

# ======================================
# ANCHOS DE COLUMNAS
# ======================================

ws.column_dimensions["A"].width = 5.3
ws.column_dimensions["B"].width = 1.3
ws.column_dimensions["C"].width = 2.5
ws.column_dimensions["D"].width = 11

ws.column_dimensions["E"].width = 29
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 18

# ======================================
# ALTURA FILAS
# ======================================

ws.row_dimensions[1].height = 21.6
ws.row_dimensions[2].height = 12.6
ws.row_dimensions[3].height = 12.6
ws.row_dimensions[4].height = 12.6

# ======================================
# VARIABLES DINÁMICAS
# ======================================

fecha = "5-11 de mayo"
lectura_biblia = "LECTURA SEMANAL DE LA BIBLIA"

presidente = "Israel Pérez"
consejero = "Juan López"
oracion_1 = "Carlos Mendoza"

hora_inicio = "19:30"
cancion_1 = "Canción 1 y oración"

# ======================================
# COMBINAR CELDAS
# ======================================

ws.merge_cells("A1:D1")
ws.merge_cells("E1:H1")

ws.merge_cells("A2:D2")
ws.merge_cells("E2:F2")

ws.merge_cells("D4:F4")

# ======================================
# FILA 1
# ======================================

ws["A1"] = "Congregación Gatazo"
ws["E1"] = "Programa para la reunión de entre semana"

ws["A1"].font = Font(
    name="Calibri",
    size=11,
    bold=True
)

ws["E1"].font = Font(
    name="Cambria",
    size=16.5,
    bold=True
)

ws["A1"].alignment = Alignment(
    horizontal="center",
    vertical="bottom"
)

ws["E1"].alignment = Alignment(
    horizontal="right",
    vertical="center"
)

borde_doble = Border(
    bottom=Side(
        style="double",
        color="000000"
    )
)

for col in range(1, 9):
    ws.cell(1, col).border = borde_doble

# ======================================
# FILA 2
# ======================================

ws["A2"] = fecha
ws["E2"] = lectura_biblia

ws["G2"] = "Presidente:"
ws["H2"] = presidente

ws["A2"].font = Font(
    name="Calibri",
    size=9,
    bold=True
)

ws["E2"].font = Font(
    name="Calibri",
    size=11,
    bold=True
)

ws["G2"].font = Font(
    name="Calibri",
    size=8,
    bold=True,
    color="575A5D"
)

ws["H2"].font = Font(
    name="Calibri",
    size=9
)

ws["A2"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

ws["E2"].alignment = Alignment(
    horizontal="left",
    vertical="center"
)

ws["G2"].alignment = Alignment(
    horizontal="right",
    vertical="center"
)

ws["H2"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

borde_derecho = Border(
    right=Side(
        style="medium",
        color="000000"
    )
)

for celda in ["A2", "B2", "C2", "D2"]:
    ws[celda].border = borde_derecho

# ======================================
# FILA 3
# ======================================

ws["G3"] = "Consejero de la Sala Auxiliar:"
ws["H3"] = consejero

ws["G3"].font = Font(
    name="Calibri",
    size=8,
    bold=True,
    color="575A5D"
)

ws["H3"].font = Font(
    name="Calibri",
    size=9
)

ws["G3"].alignment = Alignment(
    horizontal="right",
    vertical="center"
)

ws["H3"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

# ======================================
# FILA 4
# ======================================

# Datos
ws["A4"] = hora_inicio
ws["C4"] = "■"
ws["D4"] = cancion_1

ws["G4"] = "Oración:"
ws["H4"] = oracion_1

# --------------------------------------
# FUENTES
# --------------------------------------

# Hora
ws["A4"].font = Font(
    name="Calibri",
    size=9,
    color="575A5D"
)

# Cuadro
ws["C4"].font = Font(
    name="Calibri",
    size=9,
    color="575A5D"
)

# Canción
ws["D4"].font = Font(
    name="Calibri",
    size=10
)

# Etiqueta oración
ws["G4"].font = Font(
    name="Calibri",
    size=8,
    bold=True,
    color="575A5D"
)

# Valor oración
ws["H4"].font = Font(
    name="Calibri",
    size=9
)

# --------------------------------------
# ALINEACIÓN
# --------------------------------------

# Hora
ws["A4"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

# Cuadro
ws["C4"].alignment = Alignment(
    horizontal="right",
    vertical="center"
)

# Canción
ws["D4"].alignment = Alignment(
    horizontal="left",
    vertical="center"
)

# Oración
ws["G4"].alignment = Alignment(
    horizontal="right",
    vertical="center"
)

ws["H4"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

# ======================================
# FILA 5
# ======================================

ws.row_dimensions[5].height = 12.6

# --------------------------------------
# HORA (+5 MIN)
# --------------------------------------

ws["A5"] = "19:35"

# --------------------------------------
# CUADRO
# --------------------------------------

ws["C5"] = "■"

# --------------------------------------
# COMBINAR D-F
# --------------------------------------

ws.merge_cells("D5:F5")

# --------------------------------------
# TEXTO
# --------------------------------------

ws["D5"] = "Palabras de Introducción (1 min.)"

# --------------------------------------
# FUENTES
# --------------------------------------

# Hora
ws["A5"].font = Font(
    name="Calibri",
    size=9,
    color="575A5D"
)

# Cuadro
ws["C5"].font = Font(
    name="Calibri",
    size=9,
    color="575A5D"
)

# Texto principal
ws["D5"].font = Font(
    name="Calibri",
    size=10
)

# --------------------------------------
# ALINEACIÓN
# --------------------------------------

# Hora
ws["A5"].alignment = Alignment(
    horizontal="center",
    vertical="center"
)

# Cuadro
ws["C5"].alignment = Alignment(
    horizontal="right",
    vertical="center"
)

# Texto
ws["D5"].alignment = Alignment(
    horizontal="left",
    vertical="center"
)

# ======================================
# GUARDAR
# ======================================

wb.save("programa.xlsx")

print("Archivo generado: programa.xlsx")