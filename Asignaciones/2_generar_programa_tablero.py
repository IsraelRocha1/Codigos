import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
import os
import time
from win32com.client import Dispatch

# ========= RUTAS DEL PROYECTO =========
base_proyecto = os.path.dirname(os.path.abspath(__file__))

carpeta_recursos = os.path.join(base_proyecto, "Recursos")
carpeta_Tablero = os.path.join(base_proyecto, "Tablero")

# (Opcional) crear si no existen
os.makedirs(carpeta_recursos, exist_ok=True)
os.makedirs(carpeta_Tablero, exist_ok=True)

# ========= ARCHIVOS EN RECURSOS =========
archivo_asignaciones = os.path.join(carpeta_recursos, "asignaciones.xlsx")
archivo_identificadores = os.path.join(carpeta_recursos, "identificadores_valores.xlsx")

# ========= MODELOS (AJUSTA SI ESTÁN EN OTRA CARPETA) =========
# Si tus modelos están en Recursos, déjalo así.
# Si están en la raíz, cambia `carpeta_modelos = base_proyecto`.
carpeta_modelos = carpeta_recursos

modelo_por_condicion = {
    "A": os.path.join(carpeta_modelos, "modelo1.xlsx"),
    "B": os.path.join(carpeta_modelos, "modelo2.xlsx"),
    "C": os.path.join(carpeta_modelos, "modelo3.xlsx"),
    "D": os.path.join(carpeta_modelos, "modelo4.xlsx"),
}

hoja_por_condicion = {
    "A": "Hoja1",
    "B": "Hoja2",
    "C": "Hoja3",
    "D": "Hoja4",
}

columnas_html = {
    "html1": {"ids": [1, 3, 5], "valores": [2, 4, 6], "columna_clave": 0},
    "html2": {"ids": [9, 11, 13], "valores": [10, 12, 14], "columna_clave": 8},
    "html3": {"ids": [17, 19, 21], "valores": [18, 20, 22], "columna_clave": 16},
    "html4": {"ids": [25, 27, 29], "valores": [26, 28, 30], "columna_clave": 24},
}

# ========= LEER EXCELS =========
df_asig = pd.read_excel(archivo_asignaciones, header=None)
libro_ids = pd.read_excel(archivo_identificadores, sheet_name=None, header=None)

# ========= FUNCIÓN PARA PROCESAR HTML =========
def procesar_html(nombre_html, config, condicion):
    try:
        print(f"\n🔄 Procesando {nombre_html} → condición {condicion}")

        if condicion not in hoja_por_condicion:
            raise ValueError(f"Condición inválida: {condicion}")

        hoja_ids = hoja_por_condicion[condicion]
        if hoja_ids not in libro_ids:
            raise FileNotFoundError(f"No existe la hoja '{hoja_ids}' en {archivo_identificadores}")

        df_ids = libro_ids[hoja_ids]
        modelo_path = modelo_por_condicion[condicion]
        if not os.path.exists(modelo_path):
            raise FileNotFoundError(f"No se encontró el modelo: {modelo_path}")

        wb_modelo = load_workbook(modelo_path)
        if "plantilla" not in wb_modelo.sheetnames:
            raise ValueError(f"El modelo {modelo_path} no tiene hoja 'plantilla'")

        ws_modelo = wb_modelo["plantilla"]

        for i in range(len(df_asig)):
            for j in range(3):
                try:
                    id_asig = df_asig.iat[i, config["ids"][j]]
                    valor = df_asig.iat[i, config["valores"][j]]

                    id_esperado = df_ids.iat[i, j * 2]
                    celda_destino = df_ids.iat[i, j * 2 + 1]

                    if pd.notna(id_asig) and pd.notna(valor) and pd.notna(id_esperado) and pd.notna(celda_destino):
                        if str(id_asig).strip() == str(id_esperado).strip():
                            ws_modelo[str(celda_destino).strip()] = valor

                except Exception as e:
                    print(f"⚠️ Error fila {i}, bloque {j}: {e}")
                    continue

        # Guardar temporal XLSX en Tablero
        temporal_path = os.path.join(carpeta_Tablero, f"temp_{nombre_html}.xlsx")
        wb_modelo.save(temporal_path)
        time.sleep(1)

        # Exportar a PDF en Tablero
        pdf_path = os.path.join(carpeta_Tablero, f"{nombre_html}.pdf")

        excel = Dispatch("Excel.Application")
        excel.DisplayAlerts = False
        excel.Visible = False

        wb_excel = excel.Workbooks.Open(os.path.abspath(temporal_path))
        wb_excel.Worksheets("plantilla").ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        wb_excel.Close(False)
        excel.Quit()

        print(f"✅ PDF generado: {pdf_path}")
        return pdf_path

    except Exception as e:
        print(f"❌ Error al procesar {nombre_html}: {e}")
        return None


# ========= PROCESAMIENTO DE HTMLs =========
pdf_superiores = []
pdf_inferiores = []

for html, config in columnas_html.items():
    col_id = config["columna_clave"]
    fm_count = (df_asig[col_id] == "fm_row mm_part").sum()
    lac_count = (df_asig[col_id] == "lac_row mm_part").sum()

    if fm_count <= 18 and lac_count <= 5:
        condicion = "A"
    elif fm_count <= 18 and lac_count > 5:
        condicion = "B"
    elif fm_count > 18 and lac_count <= 5:
        condicion = "C"
    else:
        condicion = "D"

    print(f"\n📊 {html}: fm_row = {fm_count}, lac_row = {lac_count} → condición {condicion}")

    pdf_generado = procesar_html(html, config, condicion)
    if not pdf_generado:
        continue

    # Clasificar como superior o inferior para fusión
    if html in ("html1", "html3"):
        pdf_superiores.append(pdf_generado)
    else:
        pdf_inferiores.append(pdf_generado)


# ========= COMBINAR PARES EN UN SOLO PDF =========
def combinar_pares_de_pdfs_con_texto(
    pdf_superiores, pdf_inferiores, salida_path,
    texto="S-140-S 11/23",
    offset_y=50,  # +sube / -baja (pt)
    x_texto=20,
    y_margen_inferior=30,
    font_size=8
):
    nuevo_pdf = fitz.open()
    ancho, alto = fitz.paper_size("a4")

    pares = min(len(pdf_superiores), len(pdf_inferiores))
    if pares == 0:
        print("\n⚠️ No hay pares de PDFs para combinar.")
        return

    for k in range(pares):
        sup_path = pdf_superiores[k]
        inf_path = pdf_inferiores[k]

        pdf_sup = fitz.open(sup_path)
        pdf_inf = fitz.open(inf_path)

        pagina = nuevo_pdf.new_page(width=ancho, height=alto)

        # Superior (ocupa toda la página, como en tu código)
        pagina.show_pdf_page(fitz.Rect(0, 0, ancho, alto), pdf_sup, 0)

        # Inferior desplazado
        offset = alto / 2 - offset_y
        pagina.show_pdf_page(fitz.Rect(0, offset, ancho, offset + alto), pdf_inf, 0)

        # Texto
        y_texto = alto - y_margen_inferior
        pagina.insert_text(
            (x_texto, y_texto),
            texto,
            fontsize=font_size,
            fontname="helv",
            fill=(0, 0, 0)
        )

        pdf_sup.close()
        pdf_inf.close()

    nuevo_pdf.save(salida_path)
    nuevo_pdf.close()
    print(f"\n✅ Programa final generado: {salida_path}")


# salida en la raíz del proyecto
ruta_programa = os.path.join(base_proyecto, "programa.pdf")

combinar_pares_de_pdfs_con_texto(
    pdf_superiores=pdf_superiores,
    pdf_inferiores=pdf_inferiores,
    salida_path=ruta_programa
)