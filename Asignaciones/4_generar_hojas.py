import re
from datetime import datetime, timedelta
import unicodedata
import os
import openpyxl
import win32com.client
from PyPDF2 import PdfMerger

# =========================
# RUTAS
# =========================
base_proyecto = os.path.dirname(os.path.abspath(__file__))

carpeta_Tablero = os.path.join(base_proyecto, "Tablero")  # temp_html*.xlsx
carpeta_recursos = os.path.join(base_proyecto, "Recursos")      # aquí está hojas.xlsx
carpeta_hojas = os.path.join(base_proyecto, "Hojas")            # aquí se guardan 1..4.pdf

ruta_hojas_xlsx = os.path.join(carpeta_recursos, "hojas.xlsx")  # <- viene de Recursos
ruta_imprimir_pdf = os.path.join(base_proyecto, "imprimir.pdf") # <- raíz

# Validaciones básicas
if not os.path.exists(carpeta_Tablero):
    raise FileNotFoundError(f"No existe la carpeta Tablero: {carpeta_Tablero}")
if not os.path.exists(carpeta_recursos):
    raise FileNotFoundError(f"No existe la carpeta Recursos: {carpeta_recursos}")
if not os.path.exists(carpeta_hojas):
    raise FileNotFoundError(f"No existe la carpeta Hojas: {carpeta_hojas}")
if not os.path.exists(ruta_hojas_xlsx):
    raise FileNotFoundError(f"No se encontró hojas.xlsx en: {ruta_hojas_xlsx}")

# =========================
# DICCIONARIOS
# =========================
MESES_ES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
            7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}
MESES_NOMBRE = {v.upper(): k for k, v in MESES_ES.items()}

def normalizar(texto):
    return unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode().upper()

def obtener_dia_y_mes_jueves(texto, anio_actual=datetime.now().year):
    texto = normalizar(str(texto).strip())
    m1 = re.match(r"(\d+)-(\d+)\s+DE\s+([A-Z]+)", texto)
    m2 = re.match(r"(\d+)\s+DE\s+([A-Z]+)\s+A\s+(\d+)\s+DE\s+([A-Z]+)", texto)
    try:
        if m1:
            dia = int(m1.group(1))
            mes = MESES_NOMBRE[m1.group(3)]
        elif m2:
            dia = int(m2.group(1))
            mes = MESES_NOMBRE[m2.group(2)]
        else:
            return None, None
        fecha_lunes = datetime(anio_actual, mes, dia)
        fecha_jueves = fecha_lunes + timedelta(days=3)
        return fecha_jueves.day, MESES_ES[fecha_jueves.month]
    except Exception as e:
        print("⚠️ Error interpretando la fecha:", e)
        return None, None

def exportar_rango_a_pdf(ruta_excel, hoja, rango, ruta_pdf):
    cm_a_pulgadas = lambda cm: cm * 0.3937
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(os.path.abspath(ruta_excel))
        ws = wb.Worksheets(hoja)

        ws.PageSetup.PrintArea = rango
        ws.PageSetup.TopMargin = cm_a_pulgadas(1.5)
        ws.PageSetup.LeftMargin = cm_a_pulgadas(1.4)
        ws.PageSetup.RightMargin = cm_a_pulgadas(0.6)
        ws.PageSetup.BottomMargin = cm_a_pulgadas(0.9)

        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False
        ws.PageSetup.Orientation = 1
        ws.PageSetup.PaperSize = 9
        ws.PageSetup.CenterHorizontally = True
        ws.PageSetup.CenterVertically = True

        ws.ExportAsFixedFormat(0, os.path.abspath(ruta_pdf))
        print(f"📄 PDF generado correctamente: {ruta_pdf}")

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except:
            pass

# =========================
# PROCESAR temp_html1.xlsx a temp_html4.xlsx (en Tablero)
# =========================
for n in range(1, 5):
    ruta_temp = os.path.join(carpeta_Tablero, f"temp_html{n}.xlsx")
    if not os.path.exists(ruta_temp):
        print(f"❌ Archivo no encontrado: {ruta_temp}")
        continue

    wb_origen = openpyxl.load_workbook(ruta_temp)
    ws_origen = wb_origen.active

    texto_i2 = ws_origen["I2"].value
    if not texto_i2:
        print(f"⚠️ La celda I2 está vacía en {ruta_temp}.")
        continue

    dia, mes = obtener_dia_y_mes_jueves(texto_i2)
    contador = sum(1 for celda in ws_origen["L"] if isinstance(celda.value, str) and celda.value.strip())

    wb_destino = openpyxl.load_workbook(ruta_hojas_xlsx)

    if contador == 4:
        hoja_destino = wb_destino["Plantilla1"]
        rango_origen = ws_origen["K9:Q16"]
        fila_destino = 93
        hoja_nombre = "Plantilla1"
        hoja_destino["F87"] = dia
        hoja_destino["G87"] = mes
        rango_pdf = "A1:X76"

    elif contador == 5:
        hoja_destino = wb_destino["Plantilla2"]
        rango_origen = ws_origen["K9:Q18"]
        fila_destino = 104
        hoja_nombre = "Plantilla2"
        hoja_destino["F98"] = dia
        hoja_destino["G98"] = mes
        rango_pdf = "A1:X95"

    else:
        print(f"⚠️ Conteo inválido en {ruta_temp} (L: {contador})")
        continue

    for i, fila in enumerate(rango_origen):
        for j, celda in enumerate(fila[:7]):
            destino = hoja_destino.cell(row=fila_destino + i, column=6 + j)
            if not isinstance(destino, openpyxl.cell.cell.MergedCell):
                destino.value = celda.value

    # Guardar hojas.xlsx de vuelta en Recursos
    wb_destino.save(ruta_hojas_xlsx)

    # Guardar PDF n.pdf en Hojas
    ruta_pdf_n = os.path.join(carpeta_hojas, f"{n}.pdf")
    exportar_rango_a_pdf(ruta_hojas_xlsx, hoja_nombre, rango_pdf, ruta_pdf_n)

# =========================
# UNIR PDFs (desde Hojas) -> imprimir.pdf (en raíz)
# =========================
pdfs = [
    os.path.join(carpeta_hojas, f"{i}.pdf")
    for i in range(1, 5)
    if os.path.exists(os.path.join(carpeta_hojas, f"{i}.pdf"))
]

if pdfs:
    merger = PdfMerger()
    for pdf in pdfs:
        merger.append(pdf)
    merger.write(ruta_imprimir_pdf)
    merger.close()
    print(f"✅ Todos los PDFs han sido unidos correctamente como '{ruta_imprimir_pdf}'")
else:
    print("⚠️ No se encontraron archivos PDF para unir.")