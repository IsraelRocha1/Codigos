from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import re
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os

carpeta = r"C:\Users\israe\Downloads"

enlaces = [
    "https://wol.jw.org/es/wol/d/r4/lp-s/202026082",
    "https://wol.jw.org/es/wol/d/r4/lp-s/202026083",
    "https://wol.jw.org/es/wol/d/r4/lp-s/202026084",

    "https://wol.jw.org/es/wol/d/r4/lp-s/202026006",
    "https://wol.jw.org/es/wol/d/r4/lp-s/202026086"
]

# =========================
# CONFIGURACIÓN GENERAL
# =========================
base_proyecto = os.path.dirname(os.path.abspath(__file__))
archivo_salida = os.path.join(base_proyecto, "Recursos", "asignaciones.xlsx")
archivo_titulos = os.path.join(base_proyecto, "Recursos", "canticos_lista.xlsx")

# Puedes poner aquí todos los html que quieras
def orden_natural(nombre):
    import re
    return [int(texto) if texto.isdigit() else texto.lower()
            for texto in re.split(r'(\d+)', nombre)]

nombres = sorted(
    [f for f in os.listdir(carpeta) if f.lower().endswith(".html")],
    key=orden_natural
)
archivos_html = [os.path.join(carpeta, n) for n in nombres]
archivos_html = [os.path.join(carpeta, n) for n in nombres]

# Validación importante
if len(enlaces) != len(archivos_html):
    raise ValueError(
        f"La cantidad de enlaces ({len(enlaces)}) debe coincidir con la cantidad de archivos HTML ({len(archivos_html)})."
    )

# =========================
# FUNCIONES AUXILIARES
# =========================
def obtener_columna_inicio(html_index):
    """
    Devuelve la columna inicial del bloque de un HTML.
    html_index empieza en 1.
    
    Estructura por bloque:
    A: clase
    B: id clase
    C: asignación
    D: id asignación
    E: participante
    F: id participante
    G: tiempo
    H: libre / separación
    
    Pero en tu estructura real escribes:
    col 1 = clase
    col 3 = asignación
    col 5 = participante
    col 7 = tiempo
    
    Entonces cada bloque ocupa 8 columnas.
    """
    return 1 + (html_index - 1) * 8


def obtener_columna_canciones(html_index):
    """
    Columnas para canciones:
    html1 -> 5
    html2 -> 13
    html3 -> 21
    html4 -> 29
    patrón = 5 + (n-1)*8
    """
    return 5 + (html_index - 1) * 8


def es_nombre_valido(texto):
    if not isinstance(texto, str):
        return False
    if "(" in texto or ")" in texto:
        return False
    if any(char.isdigit() for char in texto):
        return False
    if len(texto.split()) > 4:
        return False

    texto_limpio = texto.strip().lower()
    excluidos = [
        "sala auxiliar",
        "análisis con el auditorio",
        "estudio bíblico",
        "lector",
        "oración"
    ]
    return not any(excl in texto_limpio for excl in excluidos)


def procesar_html(nombre_archivo):
    clases_en_orden = [
        "core_row mm_part",
        "tgw_row mm_part",
        "fm_row mm_part",
        "lac_row mm_part"
    ]

    with open(nombre_archivo, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    registros = {clase: [] for clase in clases_en_orden}
    tiempo_pendiente = None

    for elem in soup.find_all(["div", "p"]):
        clase = elem.get("class")
        if not clase:
            continue

        clase_str = " ".join(clase)

        if elem.name == "p" and set(clase).issuperset({
            "part-time", "text-muted", "text-end", "pe-2", "mt-1", "pt-3"
        }):
            tiempo_pendiente = elem.get_text(strip=True)

        elif clase_str in clases_en_orden:
            lineas = elem.get_text(separator="\n").strip().split("\n")
            asignacion = lineas[0].strip()
            participantes = [l.strip() for l in lineas[1:] if l.strip()]

            for participante in participantes:
                registros[clase_str].append({
                    "clase": clase_str,
                    "nombre_asignacion": asignacion,
                    "nombre_participante": participante,
                    "tiempo": tiempo_pendiente if clase_str != "core_row mm_part" else None
                })

            tiempo_pendiente = None

    return registros


# =========================
# PARTE 1: SCRAPEAR CANCIONES
# =========================
df_titulos = pd.read_excel(archivo_titulos)
mapa_titulos = dict(zip(df_titulos["Canción #"], df_titulos["Título"]))

options = webdriver.ChromeOptions()
# options.add_argument("--headless=new")

driver = webdriver.Chrome(options=options)

data = {}

for url in enlaces:
    driver.get(url)
    time.sleep(2)

    try:
        fecha = driver.find_element(By.ID, "p1").text.strip()
    except:
        fecha = "Sin fecha"

    try:
        tema = driver.find_element(By.ID, "p2").text.strip()
    except:
        tema = ""

    elementos = driver.find_elements(By.TAG_NAME, "strong")
    canciones = []
    vistos = set()

    for elem in elementos:
        texto = elem.text.strip()
        if re.match(r"^Canción\s+\d+", texto) and texto not in vistos:
            numero = " ".join(texto.split()[:2])
            titulo = mapa_titulos.get(numero, "")
            texto_final = f"{numero}: {titulo}" if titulo else texto
            canciones.append(texto_final)
            vistos.add(texto)

    canciones.insert(0, tema)
    data[fecha] = canciones

driver.quit()

# =========================
# CREAR EXCEL
# =========================
wb = Workbook()
ws = wb.active

# Escribir canciones dinámicamente
for idx, (fecha, valores) in enumerate(data.items(), start=1):
    col_idx = obtener_columna_canciones(idx)
    ws.cell(row=54, column=col_idx, value=fecha)

    for fila, valor in enumerate(valores, start=55):
        ws.cell(row=fila, column=col_idx, value=valor)

# =========================
# PARTE 2: PROCESAR HTML
# =========================
clases_en_orden = [
    "core_row mm_part",
    "tgw_row mm_part",
    "fm_row mm_part",
    "lac_row mm_part"
]

clase_dataframes = {clase: [] for clase in clases_en_orden}

for i, archivo in enumerate(archivos_html, start=1):
    registros = procesar_html(archivo)

    for clase in clases_en_orden:
        df = pd.DataFrame(registros[clase])

        # Si no hay datos, crear dataframe vacío con columnas esperadas
        if df.empty:
            df = pd.DataFrame(columns=["clase", "nombre_asignacion", "nombre_participante", "tiempo"])

        df.columns = [f"{col}_html{i}" for col in df.columns]
        clase_dataframes[clase].append(df.reset_index(drop=True))

# =========================
# ALINEAR SOLO fm_row
# =========================
bloques = []

for clase in clases_en_orden:
    lista_df = clase_dataframes[clase]

    if clase == "fm_row mm_part":
        max_filas = max(len(df) for df in lista_df) if lista_df else 0

        nueva_lista = []
        for df in lista_df:
            faltan = max_filas - len(df)
            if faltan > 0:
                vacio = pd.DataFrame({col: [None] * faltan for col in df.columns})
                df = pd.concat([df, vacio], ignore_index=True)
            nueva_lista.append(df)

        bloque = pd.concat(nueva_lista, axis=1)
    else:
        bloque = pd.concat(lista_df, axis=1)

    bloques.append(bloque)

df_html_final = pd.concat(bloques, axis=0).reset_index(drop=True)

# =========================
# LIMPIAR NOMBRES INVÁLIDOS
# =========================
for i in range(1, len(archivos_html) + 1):
    col = f"nombre_participante_html{i}"
    if col in df_html_final.columns:
        df_html_final[col] = df_html_final[col].apply(
            lambda x: x if es_nombre_valido(x) else None
        )
        
# =========================
# PEGAR DATOS EN EXCEL
# =========================
fila_inicio = 1
fila_inicio_lac = 41

for i in range(1, len(archivos_html) + 1):
    col_inicio = obtener_columna_inicio(i)

    col_clase = f"clase_html{i}"
    col_asig = f"nombre_asignacion_html{i}"
    col_part = f"nombre_participante_html{i}"
    col_tiempo = f"tiempo_html{i}"

    # detectar filas lac_row
    filas_lac = []
    filas_no_lac = []

    if col_clase in df_html_final.columns:
        for idx, valor in enumerate(df_html_final[col_clase]):
            if valor == "lac_row mm_part":
                filas_lac.append(idx)
            else:
                filas_no_lac.append(idx)

    # ---- pegar NO lac_row como siempre ----
    for pos, idx in enumerate(filas_no_lac, start=fila_inicio):
        if col_clase in df_html_final.columns:
            ws.cell(row=pos, column=col_inicio, value=df_html_final.at[idx, col_clase])

        if col_asig in df_html_final.columns:
            ws.cell(row=pos, column=col_inicio + 2, value=df_html_final.at[idx, col_asig])

        if col_part in df_html_final.columns:
            ws.cell(row=pos, column=col_inicio + 4, value=df_html_final.at[idx, col_part])

        if col_tiempo in df_html_final.columns:
            ws.cell(row=pos, column=col_inicio + 6, value=df_html_final.at[idx, col_tiempo])

    # ---- pegar lac_row desde fila 41 ----
    for pos, idx in enumerate(filas_lac, start=fila_inicio_lac):
        if col_clase in df_html_final.columns:
            ws.cell(row=pos, column=col_inicio, value=df_html_final.at[idx, col_clase])

        if col_asig in df_html_final.columns:
            ws.cell(row=pos, column=col_inicio + 2, value=df_html_final.at[idx, col_asig])

        if col_part in df_html_final.columns:
            ws.cell(row=pos, column=col_inicio + 4, value=df_html_final.at[idx, col_part])

        if col_tiempo in df_html_final.columns:
            ws.cell(row=pos, column=col_inicio + 6, value=df_html_final.at[idx, col_tiempo])


# =========================
# ASIGNAR IDs DINÁMICAMENTE
# =========================
filas_por_grupo = max(70, len(df_html_final) + fila_inicio)

# Por cada html hay 3 columnas de IDs:
# col_inicio + 1 -> ID de asignación/clase
# col_inicio + 3 -> ID de participante/asignación
# col_inicio + 5 -> ID de tiempo/participante
#
# Letras por bloque interno: A, B, C
ids_por_grupo = ["A", "B", "C"]

for i in range(1, len(archivos_html) + 1):
    col_inicio = obtener_columna_inicio(i)

    columnas_id = [
        col_inicio + 1,  # junto a clase/asignación
        col_inicio + 3,  # junto a asignación/participante
        col_inicio + 5   # junto a participante/tiempo
    ]
    for fila in range(fila_inicio, filas_por_grupo + 1):
        for j, col_id in enumerate(columnas_id):
            letra = ids_por_grupo[j]
            id_generado = f"{letra}{fila - fila_inicio + 1}"

            # Poner ID si la columna siguiente tiene contenido
            if ws.cell(row=fila, column=col_id + 1).value:
                ws.cell(row=fila, column=col_id, value=id_generado)
    # =========================
# GUARDAR
# =========================
os.makedirs(os.path.dirname(archivo_salida), exist_ok=True)
wb.save(archivo_salida)

print(f"✅ Excel generado correctamente en: {archivo_salida}")