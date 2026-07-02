from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import re
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os

carpeta = r"C:\Users\israe\Downloads"
enlaces = [
    "https://wol.jw.org/es/wol/d/r4/lp-s/202026082",
    "https://wol.jw.org/es/wol/d/r4/lp-s/202026083",
    "https://wol.jw.org/es/wol/d/r4/lp-s/202026084",
    "https://wol.jw.org/es/wol/d/r4/lp-s/202026084"
]

# Configuración
base_proyecto = os.path.dirname(os.path.abspath(__file__))
archivo_salida = os.path.join(base_proyecto, "Recursos", "asignaciones.xlsx")
archivo_titulos = os.path.join(base_proyecto, "Recursos", "canticos_lista.xlsx")

nombres = ["1.html", "2.html", "3.html", "4.html"]  # <-- aquí cambias

archivos_html = [os.path.join(carpeta, n) for n in nombres]

options = webdriver.ChromeOptions()
# opcional:
# options.add_argument("--headless=new")

driver = webdriver.Chrome(options=options)  # Selenium descarga el driver correcto (v139)

# === Parte 1: Scrapear canciones ===
df_titulos = pd.read_excel(archivo_titulos)
mapa_titulos = dict(zip(df_titulos["Canción #"], df_titulos["Título"]))


options = webdriver.ChromeOptions()
data = {}
for url in enlaces:
    driver.get(url)
    time.sleep(2)
    try:
        fecha = driver.find_element(By.ID, "p1").text.strip()
    except:
        fecha = "Sin fecha"
    try:
        tema = driver.find_element(By.ID, "p2").text.strip()
    except:
        tema = ""
    elementos = driver.find_elements(By.TAG_NAME, "strong")
    canciones = []
    vistos = set()
    for elem in elementos:
        texto = elem.text.strip()
        if re.match(r"^Canción\s+\d+", texto) and texto not in vistos:
            numero = " ".join(texto.split()[:2])
            titulo = mapa_titulos.get(numero, "")
            texto_final = f"{numero}: {titulo}" if titulo else texto
            canciones.append(texto_final)
            vistos.add(texto)
    canciones.insert(0, tema)
    data[fecha] = canciones

driver.quit()

wb = Workbook()
ws = wb.active
col_canciones = [5, 13, 21, 29]  # F, N, V, AD

for col_idx, (fecha, valores) in zip(col_canciones, data.items()):
    ws.cell(row=54, column=col_idx, value=fecha)  # Encabezado justo arriba
    for fila, valor in enumerate(valores, start=55):
        ws.cell(row=fila, column=col_idx, value=valor)

# === Parte 2: Procesar archivos HTML ===
clases_en_orden = [
    "core_row mm_part",
    "tgw_row mm_part",
    "fm_row mm_part",
    "lac_row mm_part"
]

def es_nombre_valido(texto):
    if not isinstance(texto, str):
        return False
    if "(" in texto or ")" in texto:
        return False
    if any(char.isdigit() for char in texto):
        return False
    if len(texto.split()) > 4:
        return False
    texto_limpio = texto.strip().lower()
    excluidos = [
        "sala auxiliar",
        "análisis con el auditorio",
        "estudio bíblico",
        "lector",
        "oración"
    ]
    return not any(excl in texto_limpio for excl in excluidos)

def procesar_html(nombre_archivo):
    with open(nombre_archivo, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")
    registros = {clase: [] for clase in clases_en_orden}
    tiempo_pendiente = None
    for elem in soup.find_all(["div", "p"]):
        clase = elem.get("class")
        if not clase:
            continue
        clase_str = " ".join(clase)
        if elem.name == "p" and set(clase).issuperset({
            "part-time", "text-muted", "text-end", "pe-2", "mt-1", "pt-3"
        }):
            tiempo_pendiente = elem.get_text(strip=True)
        elif clase_str in clases_en_orden:
            lineas = elem.get_text(separator="\n").strip().split("\n")
            asignacion = lineas[0].strip()
            participantes = [l.strip() for l in lineas[1:] if l.strip()]
            for participante in participantes:
                registros[clase_str].append({
                    "clase": clase_str,
                    "nombre_asignacion": asignacion,
                    "nombre_participante": participante,
                    "tiempo": tiempo_pendiente if clase_str != "core_row mm_part" else None
                })
            tiempo_pendiente = None
    return registros

# Procesar HTML
clase_dataframes = {clase: [] for clase in clases_en_orden}
for i, archivo in enumerate(archivos_html, start=1):
    registros = procesar_html(archivo)
    for clase in clases_en_orden:
        df = pd.DataFrame(registros[clase])
        df.columns = [f"{col}_html{i}" for col in df.columns]
        clase_dataframes[clase].append(df.reset_index(drop=True))

# Bloques con alineación para fm_row
bloques = []
for clase in clases_en_orden:
    lista_df = clase_dataframes[clase]
    if clase == "fm_row mm_part":
        max_filas = max(len(df) for df in lista_df)
        for i in range(len(lista_df)):
            faltan = max_filas - len(lista_df[i])
            if faltan > 0:
                vacio = pd.DataFrame({col: [None]*faltan for col in lista_df[i].columns})
                lista_df[i] = pd.concat([lista_df[i], vacio], ignore_index=True)
        bloque = pd.concat(lista_df, axis=1)
    else:
        bloque = pd.concat(lista_df, axis=1)
    bloques.append(bloque)

# Combinar bloques
df_html_final = pd.concat(bloques, axis=0)

# Limpiar nombres inválidos
for i in range(1, 5):
    col = f"nombre_participante_html{i}"
    if col in df_html_final.columns:
        df_html_final[col] = df_html_final[col].apply(lambda x: x if es_nombre_valido(x) else None)

# Definiciones
columnas_por_html = {
    1: [1, 3, 5, 7],    # A, C, E, G
    2: [9, 11, 13, 15], # I, K, M, O
    3: [17, 19, 21, 23],# Q, S, U, W
    4: [25, 27, 29, 31] # Y, AA, AC, AE
}
ids_por_html = {
    1: "F",   # columna E
    2: "N",   # columna M
    3: "V",   # columna U
    4: "AD"   # columna AC
}

# Pegado de datos desde fila 7
next_row = 1
for i in range(1, 5):
    col_clase = f"clase_html{i}"
    col_asig = f"nombre_asignacion_html{i}"
    col_part = f"nombre_participante_html{i}"
    col_tiempo = f"tiempo_html{i}"
    cols = columnas_por_html[i]

    for fila, valor in enumerate(df_html_final[col_clase], start=next_row):
        ws.cell(row=fila, column=cols[0], value=valor)

    for fila, valor in enumerate(df_html_final[col_asig], start=next_row):
        ws.cell(row=fila, column=cols[1], value=valor)

    for fila, valor in enumerate(df_html_final[col_part], start=next_row):
        ws.cell(row=fila, column=cols[2], value=valor)

    for fila, valor in enumerate(df_html_final[col_tiempo], start=next_row):
        ws.cell(row=fila, column=cols[3], value=valor)

# === Asignar ID al final para evitar sobreescritura
# Generar IDs comunes para todos los html
# Mapeo de letras para IDs por bloque html
letras_id_por_html = {
    1: "A",
    2: "B",
    3: "C",
    4: "D"
}

filas_por_grupo = 70

# Asignar IDs diferentes por bloque html y mantenerlos iguales para cada fila
# Letras por columna de asignación (indexada desde 1)
letras_por_columna_asignacion = {
    1: "A",   # columna B
    3: "B",   # columna D
    5: "C",   # columna F
    9: "D",   # columna J
    11: "E",  # columna L
    13: "F",  # columna N
    17: "G",  # columna R
    19: "H",  # columna T
    21: "I",  # columna V
    25: "J",  # columna Z
    27: "K",  # columna AB
    29: "L"   # columna AD
}

filas_por_grupo = 70

# Para cada grupo html (html1 a html4)
# Mapear columnas de ID a letras fijas
letras_por_columna = {
    0: "A",  # ID de la columna B
    2: "B",  # ID de la columna D
    4: "C",  # ID de la columna F
    8: "D",  # ID de la columna J
    10: "E", # ID de la columna L
    12: "F", # ID de la columna N
    16: "G", # ID de la columna R
    18: "H", # ID de la columna T
    20: "I", # ID de la columna V
    24: "J", # ID de la columna Z
    26: "K", # ID de la columna AB
    28: "L"  # ID de la columna AD
}

filas_por_grupo = 70

# Asignar IDs por columna
# Mapear columnas de ID a letras fijas
letras_por_columna = {
    0: "A",  # Columna B
    2: "B",  # Columna D
    4: "C",  # Columna F
    8: "D",  # Columna J
    10: "E", # Columna L
    12: "F", # Columna N
    16: "G", # Columna R
    18: "H", # Columna T
    20: "I", # Columna V
    24: "J", # Columna Z
    26: "K", # Columna AB
    28: "L"  # Columna AD
}

filas_por_grupo = 70

# Asignar IDs por columna fija
# Asignar IDs reiniciados por grupo HTML
filas_por_grupo = 70
ids_por_grupo = ["A", "B", "C"]

for i, col_idx_base in enumerate([1, 9, 17, 25]):  # B, J, R, Z (columnas de IDs de cada html)
    for fila in range(1, filas_por_grupo + 1):
        for j, letra in enumerate(ids_por_grupo):  # A, B, C
            col_idx = col_idx_base + j * 2  # B, D, F ... J, L, N ...
            id_generado = f"{letra}{fila}"
            if ws.cell(row=fila, column=col_idx + 1).value:  # asignación
                ws.cell(row=fila, column=col_idx, value=id_generado)
            if ws.cell(row=fila, column=col_idx + 2).value:  # participante
                ws.cell(row=fila, column=col_idx + 1, value=id_generado)
            if ws.cell(row=fila, column=col_idx + 3).value:  # tiempo
                ws.cell(row=fila, column=col_idx + 2, value=id_generado)

# Guardar
wb.save(archivo_salida)

print("✅ Excel generado correctamente 'asignaciones.xlsx'")
