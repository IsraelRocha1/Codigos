import fitz  # PyMuPDF
import os
import re
import unicodedata
import csv
import json
import shutil
import pandas as pd

# =========================
# CONFIGURACIÓN
# =========================

mes = "mayo"
anio_actual = "2026"

import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#            ↑ sube de scripts/ → tu_proyecto/

CSV_CONTACTOS = os.path.join(BASE_DIR, "Archivos descargados desde hourglass", "hourglass-contactlist.csv")
JSON_GRUPOS = os.path.join(BASE_DIR, "Archivos descargados desde hourglass", "grupos.json")
PDF_UNICO = os.path.join(BASE_DIR, "Archivos descargados desde hourglass", "S-21.pdf")
PDF_TOTALES = os.path.join(BASE_DIR, "Archivos descargados desde hourglass", "S-21_2026.pdf")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = r"G:\Mi unidad\Congregacion\Tarjetas de Publicador\Registros_Publicadores"

# =========================
# EXTRAER NOTAS POR MES
# =========================

def extraer_anio_servicio(texto):
    m = re.search(
        r"Año de servicio\s*(\d{4})",
        texto,
        re.IGNORECASE
    )

    if m:
        return m.group(1)

    return None

MESES = [
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto"
]

def extraer_nota_mes(texto, mes_buscar):

    lineas = [
        l.strip()
        for l in texto.splitlines()
        if l.strip()
    ]

    for i, linea in enumerate(lineas):

        linea_lower = linea.lower()

        # Buscar el mes exacto
        if linea_lower.startswith(mes_buscar.lower()):

            # Quitar nombre del mes
            contenido = linea[len(mes_buscar):].strip()

            # =========================
            # ELIMINAR SOLO NÚMEROS
            # =========================

            contenido = re.sub(r'^\d+\s*', '', contenido)

            # Si ya quedó texto útil
            if contenido and len(contenido) > 2:

                # Ignorar encabezados basura
                basura = [
                    "cursos",
                    "bíblicos",
                    "horas",
                    "precursor",
                    "auxiliar"
                ]

                if not any(b in contenido.lower() for b in basura):
                    return contenido

            # =========================
            # Revisar líneas siguientes
            # =========================

            for j in range(i + 1, min(i + 4, len(lineas))):

                siguiente = lineas[j].strip()

                siguiente_lower = siguiente.lower()

                # detener si aparece otro mes
                if siguiente_lower in MESES:
                    break

                # ignorar encabezados
                ignorar = [
                    "participación",
                    "ministerio",
                    "cursos",
                    "bíblicos",
                    "precursor",
                    "horas",
                    "notas",
                    "total"
                ]

                if any(x in siguiente_lower for x in ignorar):
                    continue

                # ignorar líneas numéricas
                if re.fullmatch(r'[\d\s]+', siguiente):
                    continue

                # texto válido
                if len(siguiente) > 2:
                    return siguiente

    return ""

# =========================
# GENERAR EXCEL 
# =========================

def generar_excel_notas(doc, contactos, mapa_grupos, mes_actual):

    resultados = []

    for page in doc:

        texto = page.get_text()

        # =========================
        # FILTRAR SOLO 2026
        # =========================

        anio = extraer_anio_servicio(texto)

        if anio != anio_actual:
            continue

        # =========================
        # EXTRAER NOTA
        # =========================

        nota = extraer_nota_mes(texto, mes_actual)

        if not nota:
            continue

        nombre_pdf = extraer_nombre_pagina(texto)

        contacto = encontrar_contacto(nombre_pdf, contactos)

        if contacto:
            fullname = contacto["fullname"]

            overseer = normalizar(contacto["overseer"])
            grupo = mapa_grupos.get(overseer, "Sin_Grupo")

        else:
            fullname = nombre_pdf
            grupo = "Sin_Grupo"

        resultados.append({
            "Nombre": fullname,
            "Grupo": grupo,
            "Año Servicio": anio,
            "Mes": mes_actual.capitalize(),
            "Nota": nota
        })

    # =========================
    # EXPORTAR
    # =========================

    if not resultados:
        print(f"No se encontraron notas para {mes_actual}")
        return

    df = pd.DataFrame(resultados)

    ruta_excel = os.path.join(
        SCRIPT_DIR,
        f"Notas_{mes_actual.capitalize()}_2026.xlsx"
    )

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    df.to_excel(ruta_excel, index=False)

    # =========================
    # AJUSTAR ANCHO COLUMNAS
    # =========================

    wb = load_workbook(ruta_excel)
    ws = wb.active

    for column_cells in ws.columns:

        max_length = 0
        column = column_cells[0].column
        letra = get_column_letter(column)

        for cell in column_cells:

            try:
                valor = str(cell.value)

                if len(valor) > max_length:
                    max_length = len(valor)

            except:
                pass

        ajuste = max_length + 4

        ws.column_dimensions[letra].width = ajuste

    # Ajustar altura automática básica
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

    wb.save(ruta_excel)

    print(f"Excel generado: {ruta_excel}")

# =========================
# NORMALIZACIÓN
# =========================

def normalizar(texto):
    texto = texto.lower()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return texto.strip()

# =========================
# LIMPIAR SOLO PARA ARCHIVO
# =========================

def limpiar_nombre_archivo(nombre):
    return re.sub(r'[\\/*?:"<>|]', '', nombre).strip()

# =========================
# PALABRAS A IGNORAR
# =========================

PALABRAS_IGNORAR = {"de", "del", "la", "los"}

def limpiar_palabras(texto):
    return [p for p in texto.split() if p not in PALABRAS_IGNORAR]

# =========================
# EXTRAER NOMBRE PDF
# =========================

def formatear_nombre(nombre):
    if "," in nombre:
        partes = nombre.split(",")
        apellido = partes[0].strip()
        resto = partes[1].strip()
        return f"{resto} {apellido}"
    return nombre

def extraer_nombre_pagina(texto):
    for l in texto.split("\n"):
        l = l.strip()
        if "nombre:" in l.lower():
            nombre = l.split(":", 1)[1].strip()
            return formatear_nombre(nombre)
    return "Desconocido"

# =========================
# CARGAR CONTACTOS
# =========================

def cargar_contactos(csv_path):
    contactos = {}

    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)

        for row in reader:
            firstname = row.get('firstname', '').strip()
            middlename = row.get('middlename', '').strip()
            lastname = row.get('lastname', '').strip()

            clave = normalizar(
                " ".join(filter(None, [lastname, firstname, middlename]))
            )

            contactos[clave] = {
                "fullname": row.get("fullname", "").strip(),
                "overseer": row.get("group_overseer", "").strip(),
                "status": row.get("status", "").strip(),
                "inactive": row.get("inactive", "").strip()
            }

    return contactos

# =========================
# CARGAR GRUPOS
# =========================

def cargar_grupos(json_path):
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)
    
    # SOLO ancianos
    ancianos = data.get("ancianos", {})

    return {normalizar(k): v for k, v in ancianos.items()}

# =========================
# MATCHING INTELIGENTE
# =========================

def encontrar_contacto(nombre_pdf, contactos):
    nombre_pdf_norm = normalizar(nombre_pdf)
    palabras_pdf = set(limpiar_palabras(nombre_pdf_norm))

    mejor_match = None
    mejor_score = 0

    for clave, data in contactos.items():
        palabras_csv = set(limpiar_palabras(clave))

        score = len(palabras_pdf.intersection(palabras_csv))

        if score > mejor_score:
            mejor_score = score
            mejor_match = data

    if mejor_score >= 2:
        return mejor_match

    return None

# =========================
# CREAR CARPETAS
# =========================

def crear_estructura():
    # Borrar carpeta completa si ya existe
    if os.path.exists(OUTPUT_DIR):
        print("🧹 Eliminando carpeta anterior...")
        shutil.rmtree(OUTPUT_DIR)

    # Crear carpeta raíz
    os.makedirs(OUTPUT_DIR, exist_ok=True)

# =========================
# GUARDAR PDF
# =========================

def guardar_persona(nombre_pdf, paginas, contactos, mapa_grupos):

    contacto = encontrar_contacto(nombre_pdf, contactos)

    if contacto:
        fullname = contacto["fullname"]
        overseer = normalizar(contacto["overseer"])
        grupo = mapa_grupos.get(overseer, "Sin_Grupo")

        status = contacto.get("status", "").strip()

        if contacto.get("inactive", "").strip() == "1":
            status = "Inactivo"

        elif status.lower() == "regular pioneer":
            status = "Precursor"

        else:
            status = "Activo"

    else:
        fullname = nombre_pdf
        grupo = "Sin_Grupo"
        status = "Desconocido"
    nombre_archivo = limpiar_nombre_archivo(fullname)

    # =========================
    # TODOS POR GRUPO
    # =========================

    carpeta = os.path.join(
        OUTPUT_DIR,
        grupo
    )
    os.makedirs(carpeta, exist_ok=True)

    # Opcional: agregar tipo al nombre

    status_archivo = limpiar_nombre_archivo(status)

    ruta = os.path.join(
        carpeta,
        f"{nombre_archivo} - {status_archivo}.pdf"
    )

    nuevo_pdf = fitz.open()

    for p in paginas:
        nuevo_pdf.insert_pdf(
            p.parent,
            from_page=p.number,
            to_page=p.number
        )

    nuevo_pdf.save(ruta)
    nuevo_pdf.close()

    print(f"Guardado: {ruta}")
    
# =========================
# PROCESAR POR RANGO
# =========================

def procesar_rango(doc, contactos, mapa_grupos):

    persona_actual = None
    paginas_persona = []

    for page in doc:

        texto = page.get_text()
        nombre = extraer_nombre_pagina(texto)

        if persona_actual is None:
            persona_actual = nombre

        if nombre != persona_actual:
            guardar_persona(
                persona_actual,
                paginas_persona,
                contactos,
                mapa_grupos
            )

            paginas_persona = []
            persona_actual = nombre

        paginas_persona.append(page)

    if paginas_persona:
        guardar_persona(
            persona_actual,
            paginas_persona,
            contactos,
            mapa_grupos
        )

# =========================
# PROCESAR TOTALES
# =========================

def detectar_tipo_totales(texto):
    t = texto.lower()

    if "publicadores" in t:
        return "Publicadores"
    elif "precursores auxiliares" in t:
        return "Precursores Auxiliares"
    elif "Precursores regulares y especiales y misioneros en el campo" in t or "precursor especial" in t:
        return "Precursores Regulares"
    elif "misionero" in t:
        return "Misioneros"
    
    return "Otros"

def procesar_totales(pdf_path):
    print("Procesando TOTALES...")

    carpeta_totales = os.path.join(OUTPUT_DIR, "Totales")
    os.makedirs(carpeta_totales, exist_ok=True)

    doc = fitz.open(pdf_path)

    grupos = {}

    for page in doc:
        texto = page.get_text()
        tipo = detectar_tipo_totales(texto)

        if tipo not in grupos:
            grupos[tipo] = []

        grupos[tipo].append(page)

    #  Guardar cada tipo en su PDF
    for tipo, paginas in grupos.items():
        ruta = os.path.join(carpeta_totales, f"{tipo}.pdf")

        nuevo_pdf = fitz.open()
        for p in paginas:
            nuevo_pdf.insert_pdf(p.parent, from_page=p.number, to_page=p.number)

        nuevo_pdf.save(ruta)
        nuevo_pdf.close()

        print(f"Guardado Totales: {ruta}")

# =========================
# MAIN
# =========================

def main():
    crear_estructura()

    print("Cargando contactos...")
    contactos = cargar_contactos(CSV_CONTACTOS)

    print("Cargando grupos...")
    mapa_grupos = cargar_grupos(JSON_GRUPOS)

    print("Abriendo PDF único...")
    doc = fitz.open(PDF_UNICO)

    print("Procesando registros...")
    procesar_rango(
        doc,
        contactos,
        mapa_grupos
    )

    print("Generando Excel de notas...")
    generar_excel_notas(
        doc,
        contactos,
        mapa_grupos,
        mes_actual = mes
    )

    print("✅ Todo listo")

if __name__ == "__main__":
    main()