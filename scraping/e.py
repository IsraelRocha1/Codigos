from openpyxl import load_workbook
from gender import guess_gender  # el archivo de arriba

wb = load_workbook(r"C:\Users\israe\Downloads\contactos.xlsx")
ws = wb.active

# Agrega la columna "Genero" en la siguiente columna libre
new_col = ws.max_column + 1
ws.cell(row=1, column=new_col, value='Genero')

for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value  # columna A = Nombre
    if name is None or str(name).strip() == '':
        continue

    genero = guess_gender(str(name))
    valor = 'h' if genero == 'H' else ('m' if genero == 'M' else None)
    ws.cell(row=row, column=new_col, value=valor)

wb.save('contactos_con_genero.xlsx')
print('Listo')