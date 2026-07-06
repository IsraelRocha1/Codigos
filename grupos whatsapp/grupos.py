"""
Crea un grupo de WhatsApp y agrega contactos automáticamente vía WhatsApp Web.

REQUISITOS:
- pip install selenium webdriver-manager openpyxl
- Chrome instalado
- Tener WhatsApp Web vinculado (vas a escanear el QR la primera vez)

IMPORTANTE:
- Esto usa WhatsApp Web de forma automatizada, lo cual NO está oficialmente
  soportado por WhatsApp y puede infringir sus Términos de Servicio.
  Úsalo bajo tu propio riesgo (posible bloqueo temporal o permanente del número).
- Solo puedes agregar directo al grupo a contactos que te tengan guardado
  (o según su config de privacidad); al resto WhatsApp les enviará invitación.
- Usa pausas largas entre cada persona para reducir el riesgo de bloqueo.
"""

import time
import unicodedata
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

# --------- CONFIGURACIÓN ---------
NOMBRE_GRUPO = "Departamento de Hidratacion"  # nombre del grupo a crear
ARCHIVO_CONTACTOS = r"C:\Users\israe\Downloads\contactos.xlsx"
COLUMNA_TELEFONO = "Telefono"      # nombre de la columna que contiene el teléfono del contacto
COLUMNA_NOMBRE = "Nombre"           # nombre de la columna que contiene el nombre (solo para el reporte de fallidos)
CODIGO_PAIS = "593"                 # se antepone si el número no lo trae ya (ajusta al tuyo)
PAUSA_ENTRE_CONTACTOS = 3            # segundos, sube esto si puedes (30+ es más seguro)
ARCHIVO_FALLIDOS = "contactos_no_agregados.xlsx"   # se genera al final con nombre + teléfono
# ----------------------------------


def iniciar_driver():
    import os
    perfil_path = os.path.join(os.getcwd(), "whatsapp_session")
    os.makedirs(perfil_path, exist_ok=True)

    options = webdriver.ChromeOptions()
    options.add_argument(f"--user-data-dir={perfil_path}")  # guarda sesión, no vuelves a escanear QR
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--remote-debugging-port=9222")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get("https://web.whatsapp.com")
    print("Escanea el QR si es la primera vez... esperando a que cargue WhatsApp Web")
    WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.XPATH, '//div[@aria-label="Lista de chats"]'))
    )
    print("WhatsApp Web listo.")
    return driver


def cerrar_modales(driver):
    """Cierra cualquier popup/modal que tape la interfaz (avisos, tips, etc.)"""
    posibles_cierres = [
        '//div[@aria-label="Cerrar" or @aria-label="Close"]',
        '//button[@aria-label="Cerrar" or @aria-label="Close"]',
        '//div[@data-testid="x-viewer-close"]',
    ]
    for xpath in posibles_cierres:
        try:
            elem = driver.find_element(By.XPATH, xpath)
            if elem.is_displayed():
                elem.click()
                time.sleep(1)
        except Exception:
            pass
    # Como último recurso, ESC cierra la mayoría de los overlays de WhatsApp Web
    try:
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        time.sleep(0.5)
    except Exception:
        pass


def click_seguro(driver, elemento):
    """Intenta click con mouse real (ActionChains); si falla, click normal; si falla, click vía JS."""
    try:
        ActionChains(driver).move_to_element(elemento).pause(0.3).click(elemento).perform()
        return
    except Exception:
        pass
    try:
        elemento.click()
        return
    except Exception:
        pass
    driver.execute_script("arguments[0].click();", elemento)


def buscar_elemento_robusto(driver, xpaths, timeout=20, descripcion="elemento"):
    """Intenta varios XPaths distintos (la UI de WhatsApp Web cambia seguido)."""
    fin = time.time() + timeout
    ultimo_error = None
    while time.time() < fin:
        for xpath in xpaths:
            try:
                elem = driver.find_element(By.XPATH, xpath)
                if elem.is_displayed():
                    return elem
            except Exception as e:
                ultimo_error = e
        time.sleep(0.5)
    # Si no se encontró nada, guarda captura para depurar
    try:
        ruta_captura = f"error_{descripcion.replace(' ', '_')}.png"
        driver.save_screenshot(ruta_captura)
        print(f"⚠ No se encontró '{descripcion}'. Captura guardada en: {ruta_captura}")
    except Exception:
        pass
    raise TimeoutException(f"No se encontró '{descripcion}' con ninguno de los selectores probados.") from ultimo_error


def crear_grupo(driver, nombre_grupo):
    cerrar_modales(driver)

    xpaths_nuevo_chat = [
        '//div[@title="Nuevo chat" or @title="New chat"]',
        '//div[@aria-label="Nuevo chat" or @aria-label="New chat"]',
        '//button[@aria-label="Nuevo chat" or @aria-label="New chat"]',
        '//span[@data-icon="new-chat-outline"]',
        '//span[@data-icon="new-chat"]',
        '//div[@data-icon="new-chat-outline"]',
    ]
    nuevo_chat = buscar_elemento_robusto(driver, xpaths_nuevo_chat, descripcion="boton nuevo chat")
    click_seguro(driver, nuevo_chat)
    time.sleep(1)

    xpaths_nuevo_grupo = [
        '//div[contains(text(),"Nuevo grupo") or contains(text(),"New group")]',
        '//span[contains(text(),"Nuevo grupo") or contains(text(),"New group")]',
        '//li[contains(.,"Nuevo grupo") or contains(.,"New group")]',
    ]

    nuevo_grupo = None
    for intento in range(4):
        try:
            nuevo_grupo = buscar_elemento_robusto(
                driver, xpaths_nuevo_grupo, timeout=4, descripcion="opcion nuevo grupo"
            )
            break
        except TimeoutException:
            print(f"  Menú no apareció, reintentando clic en 'Nuevo chat' ({intento + 1}/4)...")
            nuevo_chat = buscar_elemento_robusto(driver, xpaths_nuevo_chat, descripcion="boton nuevo chat")
            click_seguro(driver, nuevo_chat)
            time.sleep(1)

    if nuevo_grupo is None:
        raise TimeoutException("No se pudo abrir el menú de 'Nuevo grupo' tras varios intentos.")

    click_seguro(driver, nuevo_grupo)
    time.sleep(1)
    print("Grupo iniciado. Ahora agregando contactos...")


def normalizar(texto):
    """Quita tildes/acentos y pasa a minúsculas, para comparar nombres sin importar el formato."""
    texto = str(texto or "")
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto.lower().strip()


def agregar_contacto(driver, telefono_contacto, nombre_contacto=""):
    try:
        xpaths_buscador = [
            '//div[@contenteditable="true"][@data-tab]',
            '//div[@contenteditable="true"][@role="textbox"]',
            '//p[@contenteditable="true"]',
            '//div[@title="Buscar nombre o número" or @title="Search name or number"]',
            '//input[@type="text"]',
        ]
        buscador = buscar_elemento_robusto(driver, xpaths_buscador, timeout=10, descripcion="buscador de contactos")

        # Limpiar ANTES de escribir, para no arrastrar texto de la búsqueda anterior
        buscador.click()
        buscador.send_keys(Keys.CONTROL + "a")
        buscador.send_keys(Keys.DELETE)
        time.sleep(0.3)

        buscador.send_keys(telefono_contacto)
        time.sleep(2)

        # Tomamos TODAS las filas visibles (puede incluir chips de contactos ya
        # seleccionados en vueltas anteriores, que usan el mismo contenedor).
        # Filtramos por la que realmente corresponde al nombre esperado, para no
        # volver a hacer clic sobre un contacto que ya fue agregado antes.
        candidatos = []
        fin = time.time() + 8
        while time.time() < fin and not candidatos:
            candidatos = driver.find_elements(By.XPATH, '//div[@data-testid="cell-frame-container"]')
            if not candidatos:
                time.sleep(0.5)

        if not candidatos:
            raise TimeoutException(f"No se encontraron resultados para {telefono_contacto}")

        resultado = None
        nombre_norm = normalizar(nombre_contacto)
        primer_token = nombre_norm.split(",")[0].split(" ")[0] if nombre_norm else ""

        if primer_token:
            for cand in candidatos:
                try:
                    texto_fila = normalizar(cand.text)
                    if primer_token in texto_fila:
                        resultado = cand
                        break
                except Exception:
                    continue

        # Si no logramos verificar por nombre (o no había nombre en el Excel),
        # usamos el primer candidato como antes.
        if resultado is None:
            resultado = candidatos[0]

        click_seguro(driver, resultado)

        # Limpiar después de seleccionar, para dejar listo el campo para el siguiente contacto
        buscador.send_keys(Keys.CONTROL + "a")
        buscador.send_keys(Keys.DELETE)
        print(f"  ✓ Agregado: {telefono_contacto}")
        return True
    except Exception as e:
        print(f"  ✗ No se pudo agregar a {telefono_contacto}: {type(e).__name__}: {e}")
        # Intentar limpiar el buscador igual, para no arrastrar el error al siguiente contacto
        try:
            buscador.send_keys(Keys.CONTROL + "a")
            buscador.send_keys(Keys.DELETE)
        except Exception:
            pass
        return False


def finalizar_creacion_grupo(driver, nombre_grupo):
    # Flecha para continuar tras seleccionar contactos
    xpaths_siguiente = [
        '//div[@aria-label="Siguiente" or @aria-label="Next"]',
        '//span[@data-icon="arrow-forward" or @data-icon="forward-icon"]',
        '//button[@aria-label="Siguiente" or @aria-label="Next"]',
    ]
    siguiente = buscar_elemento_robusto(driver, xpaths_siguiente, descripcion="boton siguiente")
    click_seguro(driver, siguiente)
    time.sleep(2)

    xpaths_nombre_grupo = [
        '//div[@contenteditable="true"][@data-tab]',
        '//div[@contenteditable="true"][@role="textbox"]',
        '//p[@contenteditable="true"]',
        '//div[@title="Escribe el nombre del grupo" or @title="Type group subject"]',
    ]
    nombre_input = buscar_elemento_robusto(driver, xpaths_nombre_grupo, descripcion="campo nombre del grupo")
    nombre_input.click()
    nombre_input.send_keys(nombre_grupo)
    time.sleep(1)

    xpaths_crear = [
        '//div[@aria-label="Crear grupo" or @aria-label="Create group"]',
        '//span[@data-icon="checkmark" or @data-icon="checkmark-medium"]',
        '//button[@aria-label="Crear grupo" or @aria-label="Create group"]',
    ]
    crear = buscar_elemento_robusto(driver, xpaths_crear, descripcion="boton crear grupo")
    click_seguro(driver, crear)
    print(f"Grupo '{nombre_grupo}' creado.")


def limpiar_telefono(valor):
    """Limpia el número: quita espacios, guiones, el .0 de Excel, y antepone código de país si falta."""
    telefono = str(valor).strip()
    telefono = telefono.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if telefono.endswith(".0"):
        telefono = telefono[:-2]
    telefono = telefono.lstrip("+")
    if CODIGO_PAIS and not telefono.startswith(CODIGO_PAIS):
        telefono = CODIGO_PAIS + telefono
    return telefono


def leer_contactos(archivo, columna_telefono=COLUMNA_TELEFONO, columna_nombre=COLUMNA_NOMBRE):
    """Lee nombre + teléfono desde un archivo .xlsx usando openpyxl.

    Devuelve una lista de dicts: [{"nombre": ..., "telefono": ...}, ...]
    """
    wb = load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active

    filas = ws.iter_rows(values_only=True)
    encabezados = next(filas, None)
    if encabezados is None:
        raise ValueError("El archivo .xlsx está vacío.")

    encabezados = [str(h).strip() if h is not None else "" for h in encabezados]
    if columna_telefono not in encabezados:
        raise ValueError(
            f"La columna '{columna_telefono}' no existe en el archivo. "
            f"Columnas disponibles: {encabezados}"
        )
    idx_tel = encabezados.index(columna_telefono)
    idx_nom = encabezados.index(columna_nombre) if columna_nombre in encabezados else None

    contactos = []
    for fila in filas:
        if idx_tel < len(fila) and fila[idx_tel] is not None:
            telefono = limpiar_telefono(fila[idx_tel])
            if not telefono:
                continue
            nombre = ""
            if idx_nom is not None and idx_nom < len(fila) and fila[idx_nom] is not None:
                nombre = str(fila[idx_nom]).strip()
            contactos.append({"nombre": nombre, "telefono": telefono})

    wb.close()
    return contactos


def guardar_fallidos(fallidos, archivo=ARCHIVO_FALLIDOS):
    """Genera un .xlsx con los contactos (nombre + teléfono) que no se pudieron agregar."""
    if not fallidos:
        print("No hubo contactos fallidos, no se genera archivo.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "No agregados"
    ws.append(["Nombre", "Telefono"])
    for c in fallidos:
        ws.append([c["nombre"], c["telefono"]])
    wb.save(archivo)
    print(f"📄 Archivo con {len(fallidos)} contacto(s) no agregado(s) guardado en: {archivo}")


def main():
    contactos = leer_contactos(ARCHIVO_CONTACTOS)
    print(f"{len(contactos)} contactos cargados desde {ARCHIVO_CONTACTOS}")

    driver = iniciar_driver()
    crear_grupo(driver, NOMBRE_GRUPO)

    agregados = 0
    fallidos = []
    for i, contacto in enumerate(contactos, start=1):
        print(f"[{i}/{len(contactos)}] Agregando {contacto['nombre']} ({contacto['telefono']})...")
        if agregar_contacto(driver, contacto["telefono"], contacto["nombre"]):
            agregados += 1
        else:
            fallidos.append(contacto)
        time.sleep(PAUSA_ENTRE_CONTACTOS)

    finalizar_creacion_grupo(driver, NOMBRE_GRUPO)
    print(f"\nProceso terminado. {agregados}/{len(contactos)} contactos agregados.")
    print("Nota: los que no se pudieron agregar directo recibirán invitación por link.")

    guardar_fallidos(fallidos)

    input("Presiona Enter para cerrar el navegador...")
    driver.quit()


if __name__ == "__main__":
    main()