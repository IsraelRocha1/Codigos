"""
Envía un mensaje fijo por WhatsApp a una lista de contactos, vía WhatsApp Web.

REQUISITOS:
- pip install selenium webdriver-manager openpyxl
- Chrome instalado
- Tener WhatsApp Web vinculado (vas a escanear el QR la primera vez)

IMPORTANTE:
- Esto usa WhatsApp Web de forma automatizada, lo cual NO está oficialmente
  soportado por WhatsApp y puede infringir sus Términos de Servicio.
  Úsalo bajo tu propio riesgo (posible bloqueo temporal o permanente del número).
- Enviar el mismo mensaje a muchos contactos en poco tiempo aumenta el riesgo
  de que tu número sea marcado como spam. Usa pausas largas entre envíos.
- Solo puedes escribirle directo a números que tengan WhatsApp activo.
"""

import time
import unicodedata
import urllib.parse
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service

# --------- CONFIGURACIÓN ---------
ARCHIVO_CONTACTOS = r"C:\Users\israe\Documents\Libro1.xlsx"
COLUMNA_TELEFONO = "Telefono"
COLUMNA_NOMBRE = "Nombre"
CODIGO_PAIS = "593"                    # ajusta al tuyo si no es Ecuador
PAUSA_ENTRE_MENSAJES = 20              # segundos; sube esto si puedes (30+ es más seguro)
ARCHIVO_FALLIDOS = "mensajes_no_enviados.xlsx"

MENSAJE = """Querido voluntario:
Esperamos que te encuentres muy bien.
Con el fin de conocer tu disponibilidad y poder agregarte al grupo de WhatsApp del departamento de Hidratación, te pedimos que, por favor, te pongas en contacto con Israel Rocha al 0993771082.
Una vez recibamos tu mensaje, te agregaremos al grupo para mantenerte informado sobre las comunicaciones y actividades del departamento.
Agradecemos mucho tu disposición para apoyar este importante evento.
Tus hermanos DEPARTAMENTO DE HIDRATACIÓN - AI QUITO 2026"""
# ----------------------------------


def iniciar_driver():
    import os
    perfil_path = os.path.join(os.getcwd(), "whatsapp_session")
    os.makedirs(perfil_path, exist_ok=True)

    options = webdriver.ChromeOptions()
    options.add_argument(f"--user-data-dir={perfil_path}")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get("https://web.whatsapp.com")
    print("Escanea el QR si es la primera vez... esperando a que cargue WhatsApp Web")
    WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.XPATH, '//div[@aria-label="Lista de chats"]'))
    )
    print("WhatsApp Web listo.")
    return driver


def normalizar(texto):
    texto = str(texto or "")
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return texto.lower().strip()


def limpiar_telefono(valor):
    telefono = str(valor).strip()
    telefono = telefono.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if telefono.endswith(".0"):
        telefono = telefono[:-2]
    telefono = telefono.lstrip("+")
    if CODIGO_PAIS and not telefono.startswith(CODIGO_PAIS):
        telefono = CODIGO_PAIS + telefono
    return telefono


def leer_contactos(archivo, columna_telefono=COLUMNA_TELEFONO, columna_nombre=COLUMNA_NOMBRE):
    wb = load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active

    filas = ws.iter_rows(values_only=True)
    encabezados = next(filas, None)
    if encabezados is None:
        raise ValueError("El archivo .xlsx está vacío.")

    encabezados = [str(h).strip() if h is not None else "" for h in encabezados]
    if columna_telefono not in encabezados:
        raise ValueError(
            f"La columna '{columna_telefono}' no existe en el archivo. "
            f"Columnas disponibles: {encabezados}"
        )
    idx_tel = encabezados.index(columna_telefono)
    idx_nom = encabezados.index(columna_nombre) if columna_nombre in encabezados else None

    contactos = []
    for fila in filas:
        if idx_tel < len(fila) and fila[idx_tel] is not None:
            telefono = limpiar_telefono(fila[idx_tel])
            if not telefono:
                continue
            nombre = ""
            if idx_nom is not None and idx_nom < len(fila) and fila[idx_nom] is not None:
                nombre = str(fila[idx_nom]).strip()
            contactos.append({"nombre": nombre, "telefono": telefono})

    wb.close()
    return contactos


def guardar_fallidos(fallidos, archivo=ARCHIVO_FALLIDOS):
    if not fallidos:
        print("No hubo envíos fallidos, no se genera archivo.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "No enviados"
    ws.append(["Nombre", "Telefono"])
    for c in fallidos:
        ws.append([c["nombre"], c["telefono"]])
    wb.save(archivo)
    print(f"📄 Archivo con {len(fallidos)} envío(s) fallido(s) guardado en: {archivo}")


def enviar_mensaje(driver, telefono, mensaje):
    """Abre el chat directo con el número (vía URL) y envía el mensaje."""
    try:
        texto_codificado = urllib.parse.quote(mensaje)
        url = f"https://web.whatsapp.com/send?phone={telefono}&text={texto_codificado}"
        driver.get(url)

        # Esperar a que cargue el chat con el cuadro de texto listo
        xpaths_caja_texto = [
            '//div[@contenteditable="true"][@data-tab="10"]',
            '//footer//div[@contenteditable="true"]',
            '//div[@contenteditable="true"][@role="textbox"]',
        ]
        caja_texto = None
        fin = time.time() + 20
        while time.time() < fin and caja_texto is None:
            for xp in xpaths_caja_texto:
                try:
                    elem = driver.find_element(By.XPATH, xp)
                    if elem.is_displayed():
                        caja_texto = elem
                        break
                except Exception:
                    pass
            if caja_texto is None:
                time.sleep(0.5)

        if caja_texto is None:
            raise TimeoutException(
                "No cargó el chat o el número no tiene WhatsApp (revisa que sea válido)."
            )

        # El texto ya viene precargado en la caja gracias al parámetro ?text=.
        # Solo hace falta presionar Enter para enviarlo.
        time.sleep(1.5)
        caja_texto.send_keys(Keys.ENTER)
        time.sleep(1.5)
        print(f"  ✓ Enviado a: {telefono}")
        return True
    except Exception as e:
        print(f"  ✗ No se pudo enviar a {telefono}: {type(e).__name__}: {e}")
        return False


def main():
    contactos = leer_contactos(ARCHIVO_CONTACTOS)
    print(f"{len(contactos)} contactos cargados desde {ARCHIVO_CONTACTOS}")

    driver = iniciar_driver()

    enviados = 0
    fallidos = []
    for i, contacto in enumerate(contactos, start=1):
        print(f"[{i}/{len(contactos)}] Enviando a {contacto['nombre']} ({contacto['telefono']})...")
        if enviar_mensaje(driver, contacto["telefono"], MENSAJE):
            enviados += 1
        else:
            fallidos.append(contacto)
        time.sleep(PAUSA_ENTRE_MENSAJES)

    print(f"\nProceso terminado. {enviados}/{len(contactos)} mensajes enviados.")
    guardar_fallidos(fallidos)

    input("Presiona Enter para cerrar el navegador...")
    driver.quit()


if __name__ == "__main__":
    main()