import pandas as pd
import json
import os
import unicodedata

# ---------------------------
# Función de normalización
# ---------------------------
def normalizar(texto):
    if pd.isna(texto):
        return ""
    texto = texto.strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('utf-8')
    return texto

# ---------------------------
# Rutas
# ---------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

json_path = os.path.join(BASE_DIR, "Archivos descargados desde hourglass", "grupos.json")
csv_path = os.path.join(BASE_DIR, "Archivos descargados desde hourglass", "hourglass-contactlist.csv")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(SCRIPT_DIR, "grupos_clasificados.xlsx")

# ---------------------------
# Cargar JSON
# ---------------------------
with open(json_path, "r", encoding="utf-8") as f:
    data = json.load(f)

ancianos = data["ancianos"]
siervos = data["siervos"]

ancianos_por_grupo = {v: k for k, v in ancianos.items()}
siervos_por_grupo = {v: k for k, v in siervos.items()}

# ---------------------------
# Cargar CSV
# ---------------------------
df = pd.read_csv(csv_path)

df["group_overseer"] = df["group_overseer"].astype(str).str.strip()
df["group_overseer"] = df["group_overseer"].replace(r'^\s*$', pd.NA, regex=True)
df["fullname"] = df["fullname"].astype(str).str.strip()

df["inactive"] = df["inactive"].astype(str).str.strip().str.lower()
df = df[~df["inactive"].isin(["1", "1.0", "true", "yes"])]

df["group_overseer_norm"] = df["group_overseer"].apply(normalizar)
df["fullname_norm"] = df["fullname"].apply(normalizar)

# ---------------------------
# PERSONAS SIN GRUPO
# ---------------------------

ancianos_norm_set = set(normalizar(x) for x in ancianos_por_grupo.keys())

sin_grupo = df[
    df["group_overseer"].isna() |
    (df["group_overseer_norm"] == "")
]["fullname"].dropna().unique()


for nombre in sin_grupo:
    print("\n⚠️ PERSONAS SIN GRUPO ASIGNADO:")
    print("-", nombre)

# ---------------------------
# Procesamiento
# ---------------------------
resultado = {}

for grupo in ancianos_por_grupo.keys():
    
    anciano = ancianos_por_grupo.get(grupo, "")
    siervo = siervos_por_grupo.get(grupo, "")
    
    anciano_norm = normalizar(anciano)
    siervo_norm = normalizar(siervo)

    miembros_df = df[df["group_overseer_norm"] == anciano_norm]

    miembros = miembros_df[
        (miembros_df["fullname_norm"] != anciano_norm) &
        (miembros_df["fullname_norm"] != siervo_norm)
    ]["fullname"].dropna().tolist()

    lista = []

    if anciano:
        lista.append(f"Anciano: {anciano}")
    if siervo:
        lista.append(f"Siervo: {siervo}")

    lista.extend(miembros)

    resultado[grupo] = lista

# ---------------------------
# EXCEL
# ---------------------------
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Grupos"

grupos_ordenados = sorted(resultado.keys(), key=lambda x: int(x.split()[-1]))
num_grupos = len(grupos_ordenados)

# ---------------------------
# ESTILOS
# ---------------------------
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

fill_title = PatternFill("solid", fgColor="1F4E79")   # azul oscuro
fill_header = PatternFill("solid", fgColor="D9E1F2")  # azul claro


font_title = Font(bold=True, size=18, color="FFFFFF")
font_header = Font(bold=True)
font_normal = Font()

center = Alignment(horizontal="center", vertical="center")

# ---------------------------
# TÍTULO
# ---------------------------
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_grupos)

cell = ws.cell(row=1, column=1)
cell.value = "Grupos para el servicio del campo"
cell.font = font_title
cell.fill = fill_title
cell.alignment = center

ws.row_dimensions[1].height = 30

# ---------------------------
# FILA VACÍA
# ---------------------------

# ---------------------------
# GRUPOS
# ---------------------------
for col, grupo in enumerate(grupos_ordenados, start=1):
    cell = ws.cell(row=3, column=col)
    cell.value = grupo
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = center

# ---------------------------
# SUPERINTENDENTE
# ---------------------------
ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=num_grupos)

cell = ws.cell(row=4, column=1)
cell.value = "Superintendente"
cell.font = font_header
cell.fill = fill_header
cell.alignment = center

# ---------------------------
# ANCIANOS
# ---------------------------
for col, grupo in enumerate(grupos_ordenados, start=1):
    anciano = ""

    for item in resultado[grupo]:
        if item.startswith("Anciano:"):
            anciano = item.replace("Anciano: ", "")

    ws.cell(row=5, column=col).value = anciano

# ---------------------------
# AUXILIAR
# ---------------------------
ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=num_grupos)

cell = ws.cell(row=6, column=1)
cell.value = "Auxiliar"
cell.font = font_header
cell.fill = fill_header
cell.alignment = center

# ---------------------------
# SIERVOS
# ---------------------------
for col, grupo in enumerate(grupos_ordenados, start=1):
    siervo = ""

    for item in resultado[grupo]:
        if item.startswith("Siervo:"):
            siervo = item.replace("Siervo: ", "")

    ws.cell(row=7, column=col).value = siervo

# ---------------------------
# SEPARACIÓN VISUAL
# ---------------------------

# ---------------------------
# MIEMBROS
# ---------------------------
fila_inicio = 9

max_filas = 0

for col, grupo in enumerate(grupos_ordenados, start=1):

    miembros = [
        item for item in resultado[grupo]
        if not item.startswith("Anciano:") and not item.startswith("Siervo:")
    ]

    max_filas = max(max_filas, len(miembros))

    for i, miembro in enumerate(miembros):
        ws.cell(row=fila_inicio + i, column=col).value = miembro

# ---------------------------
# BORDES A TODA LA TABLA (INCLUYE VACÍOS)
# ---------------------------
max_row = fila_inicio + max_filas

for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=num_grupos):
    for cell in row:
        cell.border = thin_border
        cell.alignment = center

# ---------------------------
# AUTOAJUSTE COLUMNAS
# ---------------------------
for col_idx, col in enumerate(ws.iter_cols(min_row=3), start=1):
    max_length = 0
    col_letter = get_column_letter(col_idx)

    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[col_letter].width = (max_length ) 

# ---------------------------
# CONGELAR PANELES
# ---------------------------
ws.freeze_panes = "A3"  # congela título + encabezados

# ---------------------------
# CONFIGURACIÓN DE IMPRESIÓN (PDF)
# ---------------------------
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

ws.print_area = f"A1:{get_column_letter(num_grupos)}{fila_inicio + max_filas}"

# ---------------------------
# GUARDAR
# ---------------------------
wb.save(output_path)

print(f"\n✅ Archivo profesional creado en: {output_path}")